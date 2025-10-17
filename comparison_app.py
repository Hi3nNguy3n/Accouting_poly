# Copyright by hiennm22, LocTH5 - BM UDPM
import streamlit as st
import pandas as pd
import os
import glob
import zipfile
import io
import re
import base64
import json
import os.path
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.application import MIMEApplication
from email.utils import formataddr
from google.auth.transport.requests import Request
from google.oauth2.credentials import Credentials
from google_auth_oauthlib.flow import Flow
from googleapiclient.discovery import build
from googleapiclient.errors import HttpError

# Cố gắng import pypdf và hướng dẫn cài đặt nếu thiếu
try:
    import pypdf
except ImportError:
    st.error("Thư viện pypdf chưa được cài đặt. Vui lòng chạy lệnh sau trong terminal: pip install pypdf")
    st.stop()

# --- HÀM HỖ TRỢ OAUTH2 & ĐĂNG NHẬP ---
SCOPES = [
    'https://www.googleapis.com/auth/gmail.send',
    'https://www.googleapis.com/auth/gmail.readonly',
    'openid',
    'https://www.googleapis.com/auth/userinfo.email',
    'https://www.googleapis.com/auth/userinfo.profile'
]
TOKEN_FILE = "token.json"
CREDENTIALS_FILE = "credentials.json"

def get_google_credentials(credentials_json_content):
    """Hàm này xử lý cả việc lấy credentials và thông tin người dùng."""
    creds = None
    # Tải credentials từ session nếu có
    if 'credentials' in st.session_state:
        creds_json = json.loads(st.session_state['credentials'])
        creds = Credentials.from_authorized_user_info(creds_json, SCOPES)

    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            try:
                creds.refresh(Request())
                st.session_state['credentials'] = creds.to_json()
            except Exception as e:
                st.error(f"Lỗi khi làm mới token: {e}")
                # Xóa token hỏng và yêu cầu đăng nhập lại
                if 'credentials' in st.session_state:
                    del st.session_state['credentials']
                if 'user_info' in st.session_state:
                    del st.session_state['user_info']
                return None, None
        else:
            # If there are no valid credentials, just return None.
            # The login page will handle the interactive login.
            return None, None

    # Sau khi có credentials, lấy thông tin người dùng
    try:
        service = build('oauth2', 'v2', credentials=creds)
        user_info = service.userinfo().get().execute()
        
        # --- KIỂM TRA EMAIL ---
        email = user_info.get('email', '').lower()
        if not email.endswith('@fpt.edu.vn'):
            st.error("Truy cập bị từ chối. Chỉ các tài khoản email FPT (@fpt.edu.vn) mới được phép đăng nhập.")
            # Xóa thông tin đăng nhập không hợp lệ
            keys_to_delete = ['credentials', 'user_info']
            for key in keys_to_delete:
                if key in st.session_state:
                    del st.session_state[key]
            return None, None
        # --- KẾT THÚC KIỂM TRA EMAIL ---

        st.session_state['user_info'] = user_info
        return creds, user_info
    except HttpError as error:
        st.error(f"Lỗi khi lấy thông tin người dùng: {error}")
        return None, None

def show_login_page():
    """Hiển thị trang đăng nhập và xử lý việc nhấn nút."""
    st.set_page_config(page_title="Đăng nhập", layout="centered", page_icon="🔑")
    st.title("🔑 Đăng nhập đi cục dàng!")
    st.write("Vui lòng đăng nhập bằng tài khoản Google @fpt.edu.vn của bạn để tiếp tục.")

    # Tải credentials.json
    credentials_json_content = None
    if os.path.exists(CREDENTIALS_FILE):
        with open(CREDENTIALS_FILE, 'r', encoding='utf-8') as f:
            credentials_json_content = f.read()
    else:
        st.error(f"Không tìm thấy file `{CREDENTIALS_FILE}`. Vui lòng đảm bảo file này tồn tại trong thư mục.")
        st.stop()

    try:
        # Use Flow for web applications, not InstalledAppFlow
        flow = Flow.from_client_config(
            json.loads(credentials_json_content),
            scopes=SCOPES,
            redirect_uri='https://mergegrab.streamlit.app/'  # Must match GCP Console
        )

        # Check if the user has been redirected back from Google
        query_params = st.query_params
        if 'code' in query_params:
            with st.spinner("Đang xác thực, vui lòng chờ..."):
                try:
                    # Exchange the code for credentials
                    flow.fetch_token(code=query_params.get('code'))
                    creds = flow.credentials
                    st.session_state['credentials'] = creds.to_json()

                    # Clear the query parameters from the URL
                    st.query_params.clear()

                    # Get user info and validate
                    service = build('oauth2', 'v2', credentials=creds)
                    user_info = service.userinfo().get().execute()
                    email = user_info.get('email', '').lower()

                    if not email.endswith('@fpt.edu.vn'):
                        st.error("Truy cập bị từ chối. Chỉ các tài khoản email FPT (@fpt.edu.vn) mới được phép đăng nhập.")
                        # Clear invalid session
                        keys_to_delete = ['credentials', 'user_info']
                        for key in keys_to_delete:
                            if key in st.session_state:
                                del st.session_state[key]
                    else:
                        st.session_state['user_info'] = user_info
                        st.success(f"Đăng nhập thành công! Xin chào, {user_info.get('name', 'bạn')}.")
                        # Rerun to show the main app
                        st.rerun()

                except Exception as e:
                    st.error(f"Lỗi khi xác thực với Google: {e}")
        else:
            # Show the login button
            auth_url, _ = flow.authorization_url(prompt='consent')
            st.link_button("Đăng nhập với Google", auth_url, use_container_width=True, help="Bạn sẽ được chuyển đến trang đăng nhập của Google")

    except Exception as e:
        st.error(f"Lỗi khi khởi tạo quy trình xác thực: {e}")
        st.exception(e)

def main_app():
    """Hàm chứa toàn bộ giao diện và logic của ứng dụng chính."""
    st.set_page_config(page_title="Đối chiếu FPT", layout="wide", page_icon="📊")

    # --- SIDEBAR & ĐĂNG XUẤT ---
    with st.sidebar:
        user_info = st.session_state.get('user_info', {})
        st.markdown(f"Xin chào, **{user_info.get('name', 'người dùng')}**")
        st.caption(user_info.get('email'))
        if st.button("Đăng xuất"):
            # Xóa thông tin đăng nhập khỏi session
            keys_to_delete = ['credentials', 'user_info']
            for key in keys_to_delete:
                if key in st.session_state:
                    del st.session_state[key]
            st.rerun()

    st.title("📊 Đối chiếu dữ liệu Grab & Báo cáo PDF")
    st.write("Tải lên các tệp của bạn để bắt đầu đối chiếu và xử lý.")
    st.caption("Copyright by LocTH5, Hiennm22 - BM UDPM")

    # (Dán toàn bộ phần còn lại của ứng dụng vào đây)
    # --- GIAO DIỆN NHẬP LIỆU ---
    with st.container(border=True):
        st.subheader("Tải lên các file cần thiết")
        col1, col2, col3 = st.columns(3)
        file_types = ["csv", "xls", "xlsx"]
        with col1:
            uploaded_transport_file = st.file_uploader("1. File Transport", type=file_types)
            uploaded_express_file = st.file_uploader("2. File Express", type=file_types)
        with col2:
            uploaded_invoice_file = st.file_uploader("3. File Hóa đơn", type=file_types)
            uploaded_zip_file = st.file_uploader("4. Folder Báo cáo (.zip)", type=["zip"])
        with col3:
            uploaded_xml_zip_file = st.file_uploader("5. Folder XML (.zip)", type=["zip"])

    # --- CẤU HÌNH OAUTH 2.0 ---
    # The app now automatically loads 'credentials.json' from the local directory.
    CREDENTIALS_FILE = "credentials.json"
    st.session_state.credentials_loaded = False
    if os.path.exists(CREDENTIALS_FILE):
        with open(CREDENTIALS_FILE, 'r', encoding='utf-8') as f:
            st.session_state.credentials_json_content = f.read()
        st.session_state.credentials_loaded = True

    # --- BẮT ĐẦU XỬ LÝ KHI CÓ ĐỦ FILE ---
    if (uploaded_transport_file is not None or uploaded_express_file is not None) and uploaded_invoice_file is not None:
        try:
            employee_to_unit_map, unit_to_email_map = load_mapping_data()

            # --- 1. ĐỌC VÀ LÀM SẠCH DỮ LIỆU GỐC ---
            source_dfs = []
            if uploaded_transport_file:
                try:
                    if uploaded_transport_file.name.endswith('.csv'):
                        df_transport_single = pd.read_csv(uploaded_transport_file, skiprows=7)
                    else:
                        df_transport_single = pd.read_excel(uploaded_transport_file, skiprows=7)
                    
                    if df_transport_single.shape[1] > 10:
                        # Rename Booking ID from Column K (index 10)
                        df_transport_single.rename(columns={df_transport_single.columns[10]: 'Booking ID'}, inplace=True)
                        # Rename Employee Name from Column C (index 2)
                        df_transport_single.rename(columns={df_transport_single.columns[2]: 'Employee Name'}, inplace=True)
                        source_dfs.append(df_transport_single)
                    else:
                        st.warning(f"File Transport '{uploaded_transport_file.name}' dường như không hợp lệ (cần ít nhất 11 cột). Bỏ qua file này.")
                except Exception as e:
                    st.error(f"Lỗi khi đọc file Transport '{uploaded_transport_file.name}': {e}")

            if uploaded_express_file:
                try:
                    if uploaded_express_file.name.endswith('.csv'):
                        df_express_single = pd.read_csv(uploaded_express_file, skiprows=7)
                    else:
                        df_express_single = pd.read_excel(uploaded_express_file, skiprows=7)

                    if df_express_single.shape[1] > 9:
                        # Rename Booking ID from Column J (index 9)
                        df_express_single.rename(columns={df_express_single.columns[9]: 'Booking ID'}, inplace=True)
                        # Rename Employee Name from Column C (index 2)
                        df_express_single.rename(columns={df_express_single.columns[2]: 'Employee Name'}, inplace=True)
                        source_dfs.append(df_express_single)
                    else:
                        st.warning(f"File Express '{uploaded_express_file.name}' dường như không hợp lệ (cần ít nhất 10 cột). Bỏ qua file này.")
                except Exception as e:
                    st.error(f"Lỗi khi đọc file Express '{uploaded_express_file.name}': {e}")

            if not source_dfs:
                st.error("Không thể xử lý file Transport hoặc Express. Vui lòng kiểm tra lại định dạng file.")
                st.stop()
            
            df_transport = pd.concat(source_dfs, ignore_index=True)

            if uploaded_invoice_file.name.endswith('.csv'):
                df_invoice = pd.read_csv(uploaded_invoice_file)
            elif uploaded_invoice_file.name.endswith('.xls'):
                try:
                    df_invoice = pd.read_html(uploaded_invoice_file)[0]
                except Exception:
                    df_invoice = pd.read_excel(uploaded_invoice_file, engine='xlrd')
            else:
                df_invoice = pd.read_excel(uploaded_invoice_file)

            df_transport.columns = df_transport.columns.str.strip()
            df_invoice.columns = df_invoice.columns.str.strip()

            if len(df_transport.columns) > 9:
                pickup_col_name = df_transport.columns[7]
                dropoff_col_name = df_transport.columns[9]
                df_transport.rename(columns={
                    pickup_col_name: 'GEMINI_PICKUP_ADDRESS',
                    dropoff_col_name: 'GEMINI_DROPOFF_ADDRESS'
                }, inplace=True)

            if df_invoice.shape[1] < 13:
                st.error(f"File Hóa đơn không có đủ 13 cột. Chỉ tìm thấy {df_invoice.shape[1]} cột.")
                st.stop()
            rename_dict = {
                df_invoice.columns[1]: 'pdf_link_key',
                df_invoice.columns[12]: 'summary_ma_nhan_hoa_don'
            }
            if len(df_invoice.columns) > 4:
                rename_dict[df_invoice.columns[4]] = 'GEMINI_NGAY_HD_INVOICE'
                rename_dict[df_invoice.columns[5]] = 'HINH_THUC_TT'
                rename_dict[df_invoice.columns[6]] = 'TIEN_TRC_THUE'
                rename_dict[df_invoice.columns[7]] = 'TIEN_THUE8'
                rename_dict[df_invoice.columns[8]] = 'TONG_TIEN'
                rename_dict[df_invoice.columns[15]] = 'NGAY_BOOKING'
                rename_dict[df_invoice.columns[16]] = 'SO_HOA_DON'
            df_invoice.rename(columns=rename_dict, inplace=True)

            # Check for unmatched Booking IDs between Transport and Invoice files
            transport_ids = set(df_transport['Booking ID'].dropna())
            invoice_ids = set(df_invoice['Booking'].dropna())

            unmatched_transport_ids = transport_ids - invoice_ids
            unmatched_invoice_ids = invoice_ids - transport_ids

            if unmatched_transport_ids:
                st.warning(f"Tìm thấy {len(unmatched_transport_ids)} Booking ID chỉ có trong file Transport (không có trong file Hóa đơn):")
                with st.expander("Xem danh sách ID bị thừa từ file Transport"):
                    st.dataframe(pd.DataFrame(sorted(list(unmatched_transport_ids)), columns=["Booking ID"]))

            if unmatched_invoice_ids:
                st.warning(f"Tìm thấy {len(unmatched_invoice_ids)} Booking ID chỉ có trong file Hóa đơn (không có trong file Transport):")
                with st.expander("Xem danh sách ID bị thừa từ file Hóa đơn"):
                    st.dataframe(pd.DataFrame(sorted(list(unmatched_invoice_ids)), columns=["Booking ID"]))


            matching_ids = list(set(df_transport['Booking ID'].dropna()) & set(df_invoice['Booking'].dropna()))
            if not matching_ids:
                st.warning("Không tìm thấy Booking ID nào trùng khớp giữa hai file đầu vào.")
                st.stop()

            df_merged = pd.merge(df_transport[df_transport['Booking ID'].isin(matching_ids)], df_invoice[df_invoice['Booking'].isin(matching_ids)], left_on='Booking ID', right_on='Booking', suffixes=('_transport', '_invoice'))
            df_merged['Đơn vị'] = df_merged['Employee Name'].map(employee_to_unit_map)
            df_merged['Đơn vị'].fillna('Không xác định', inplace=True)

            # --- 3. XỬ LÝ FOLDER PDF (NẾU CÓ) ---
            count_no_pdf = 0
            if uploaded_zip_file is not None:
                pdf_data = []
                with zipfile.ZipFile(uploaded_zip_file, 'r') as zip_ref:
                    pdf_file_names = [name for name in zip_ref.namelist() if name.lower().endswith('.pdf') and not name.startswith('__MACOSX')]
                    st.info(f"Bắt đầu xử lý {len(pdf_file_names)} file PDF từ tệp .zip...")
                    progress_bar = st.progress(0, text="Đang xử lý file PDF...")
                    for i, filename in enumerate(pdf_file_names):
                        try:
                            key_from_filename = os.path.basename(filename).split('_')[2]
                            with zip_ref.open(filename) as pdf_file:
                                pdf_content = pdf_file.read()
                                pdf_reader = pypdf.PdfReader(io.BytesIO(pdf_content))
                                text = "".join([page.extract_text() or "" for page in pdf_reader.pages])
                                found_code = "Không tìm thấy trong PDF"
                                if "Mã nhận hóa đơn" in text:
                                    parts = text.split("Mã nhận hóa đơn")
                                    if len(parts) > 1:
                                        code = parts[1].split('\n')[0].replace(":", "").strip()
                                        if code: found_code = code
                                ngay_hd_str = "Không tìm thấy"
                                match = re.search(r'Ngày\s*(\d{1,2})\s*tháng\s*(\d{1,2})\s*năm\s*(\d{4})', text, re.IGNORECASE)
                                if match:
                                    day, month, year = match.groups()
                                    ngay_hd_str = f"{day.zfill(2)}/{month.zfill(2)}/{year}"
                                pdf_data.append({'pdf_link_key_str': key_from_filename, 'Mã hóa đơn từ PDF': found_code, 'Ngay_HD_pdf': ngay_hd_str, 'pdf_content': pdf_content, 'pdf_filename': os.path.basename(filename)})
                        except Exception as e:
                            st.warning(f"Lỗi khi đọc file {filename} trong zip: {e}")
                        progress_bar.progress((i + 1) / len(pdf_file_names), text=f"Đang xử lý: {os.path.basename(filename)}")
                if pdf_data:
                    df_pdf_data = pd.DataFrame(pdf_data)
                    df_merged['pdf_link_key_str'] = df_merged['pdf_link_key'].astype(str)

                    # Check for PDF files that don't match any invoice row
                    merged_keys = set(df_merged['pdf_link_key_str'])
                    pdf_keys = set(df_pdf_data['pdf_link_key_str'])
                    unmatched_pdf_keys = pdf_keys - merged_keys

                    if unmatched_pdf_keys:
                        unmatched_pdf_files = df_pdf_data[df_pdf_data['pdf_link_key_str'].isin(unmatched_pdf_keys)]
                        st.warning(f"Tìm thấy {len(unmatched_pdf_files)} file PDF không khớp với bất kỳ dòng nào trong file Hóa đơn:")
                        with st.expander("Xem danh sách file PDF bị thừa"):
                            st.dataframe(unmatched_pdf_files[['pdf_filename', 'pdf_link_key_str']])

                    df_merged = pd.merge(df_merged, df_pdf_data, on='pdf_link_key_str', how='left')
                    count_no_pdf = df_merged['pdf_filename'].isnull().sum()

            # --- 3.1. XỬ LÝ FOLDER XML (NẾU CÓ) ---
            if uploaded_xml_zip_file is not None:
                xml_data = []
                with zipfile.ZipFile(uploaded_xml_zip_file, 'r') as zip_ref:
                    xml_file_names = [name for name in zip_ref.namelist() if name.lower().endswith('.xml') and not name.startswith('__MACOSX')]
                    st.info(f"Bắt đầu xử lý {len(xml_file_names)} file XML từ tệp .zip...")
                    progress_bar_xml = st.progress(0, text="Đang xử lý file XML...")
                    for i, filename in enumerate(xml_file_names):
                        try:
                            base = os.path.basename(filename)
                            key_from_filename = ''
                            try:
                                # Attempt 1: PDF-style naming (e.g., 1_C25MGA_2565515_....xml)
                                key_from_filename = base.split('_')[2]
                            except IndexError:
                                # Attempt 2: Simple naming (e.g., 2565515.xml)
                                key_from_filename = base.split('.')[0]

                            if not key_from_filename.strip():
                                st.warning(f"Không thể lấy key từ tên file XML: '{filename}'. Bỏ qua file này.")
                                continue

                            with zip_ref.open(filename) as xml_file:
                                xml_content = xml_file.read()
                                text = xml_content.decode('utf-8', errors='ignore')

                                # Extract invoice code from XML using regex on common tags
                                found_code = "Không tìm thấy trong XML"
                                code_match = re.search(r'<InvoiceCode>(.*?)</InvoiceCode>', text, re.IGNORECASE) or \
                                             re.search(r'<MaNhanHoaDon>(.*?)</MaNhanHoaDon>', text, re.IGNORECASE) or \
                                             re.search(r'<TransactionID>(.*?)</TransactionID>', text, re.IGNORECASE) or \
                                             re.search(r'<Fkey>(.*?)</Fkey>', text, re.IGNORECASE) # Another common one
                                if code_match:
                                    found_code = code_match.group(1).strip()

                                # Extract invoice date from XML using regex
                                ngay_hd_str = "Không tìm thấy"
                                date_match = re.search(r'<IssuedDate>(.*?)</IssuedDate>', text, re.IGNORECASE) # YYYY-MM-DD
                                if date_match:
                                    try:
                                        # Handle various date formats that might be in the tag
                                        dt = pd.to_datetime(date_match.group(1))
                                        ngay_hd_str = dt.strftime('%d/%m/%Y')
                                    except Exception:
                                        # If parsing fails, try to use the raw value
                                        ngay_hd_str = date_match.group(1).strip()
                                else:
                                    # Fallback to the same regex as for PDFs
                                    match = re.search(r'Ngày\s*(\d{1,2})\s*tháng\s*(\d{1,2})\s*năm\s*(\d{4})', text, re.IGNORECASE)
                                    if match:
                                        day, month, year = match.groups()
                                        ngay_hd_str = f"{day.zfill(2)}/{month.zfill(2)}/{year}"

                                xml_data.append({
                                    'pdf_link_key_str': key_from_filename,
                                    'Mã hóa đơn từ XML': found_code,
                                    'Ngay_HD_xml': ngay_hd_str,
                                    'xml_content': xml_content,
                                    'xml_filename': os.path.basename(filename)
                                })
                        except Exception as e:
                            st.warning(f"Lỗi khi đọc file XML {filename} trong zip: {e}")
                        progress_bar_xml.progress((i + 1) / len(xml_file_names), text=f"Đang xử lý: {os.path.basename(filename)}")
                
                if xml_data:
                    df_xml_data = pd.DataFrame(xml_data)
                    if 'pdf_link_key_str' not in df_merged.columns:
                        df_merged['pdf_link_key_str'] = df_merged['pdf_link_key'].astype(str)
                    
                    # Check for XML files that don't match any invoice row
                    merged_keys_for_xml = set(df_merged['pdf_link_key_str'])
                    xml_keys = set(df_xml_data['pdf_link_key_str'])
                    unmatched_xml_keys = xml_keys - merged_keys_for_xml

                    if unmatched_xml_keys:
                        unmatched_xml_files = df_xml_data[df_xml_data['pdf_link_key_str'].isin(unmatched_xml_keys)]
                        st.warning(f"Tìm thấy {len(unmatched_xml_files)} file XML không khớp với bất kỳ dòng nào trong file Hóa đơn:")
                        with st.expander("Xem danh sách file XML bị thừa"):
                            st.dataframe(unmatched_xml_files[['xml_filename', 'pdf_link_key_str']])

                    df_merged = pd.merge(df_merged, df_xml_data, on='pdf_link_key_str', how='left')

            # --- 4. THỐNG KÊ VÀ HIỂN THỊ ---
            if count_no_pdf > 0:
                st.warning(f"### ⚠️ Chú ý: Có {count_no_pdf} hóa đơn không có file PDF tương ứng.")
                with st.expander("Xem danh sách và thống kê các hóa đơn thiếu PDF"):
                    df_missing_pdfs = df_merged[df_merged['pdf_filename'].isnull()]
                    
                    # 1. Display statistics table
                    st.markdown("#### Thống kê theo Đơn vị")
                    missing_stats = df_missing_pdfs.groupby('Đơn vị').agg(
                        so_hoa_don_thieu=('Booking ID', 'count'),
                        tong_tien_thieu=('Total Fare', 'sum')
                    ).reset_index()
                    missing_stats.rename(columns={
                        'so_hoa_don_thieu': 'Số hóa đơn thiếu PDF',
                        'tong_tien_thieu': 'Tổng tiền (ước tính)'
                    }, inplace=True)
                    st.dataframe(missing_stats)

                    # 2. Display the raw list
                    st.markdown("---")
                    st.markdown("#### Danh sách chi tiết")
                    
                    cols_to_show = ['Booking ID', 'Employee Name', 'Đơn vị', 'pdf_link_key', 'Total Fare']
                    date_cols = ['Date', 'Date of Trip', 'Trip Date', 'Ngày', 'Date & Time (GMT+7)']
                    available_date_col = find_col(df_missing_pdfs, date_cols)
                    if available_date_col:
                        cols_to_show.append(available_date_col)

                    existing_cols_to_show = [col for col in cols_to_show if col in df_missing_pdfs.columns]
                    st.dataframe(df_missing_pdfs[existing_cols_to_show])

            st.header("Kết quả đối chiếu")

            with st.container(border=True):
                st.subheader("Bảng thống kê tổng hợp")

                if 'Employee Name' not in df_merged.columns:
                    st.error("Không thể tạo bảng thống kê: Thiếu cột 'Employee Name' trong dữ liệu đã hợp nhất.")
                else:
                    agg_dict = {}
                    if 'Booking ID' in df_merged.columns:
                        agg_dict['So chuyen'] = ('Booking ID', 'count')
                    if 'Total Fare' in df_merged.columns:
                        agg_dict['Tong tien (VND)'] = ('Total Fare', 'sum')

                    if not agg_dict:
                        st.warning("Không thể tạo bảng thống kê: Thiếu cả cột 'Booking ID' và 'Total Fare'.")
                    else:
                        summary_df = df_merged.groupby('Employee Name').agg(**agg_dict).reset_index()

                        if 'Tong tien (VND)' in summary_df.columns:
                            summary_df = summary_df.sort_values('Tong tien (VND)', ascending=False).reset_index(drop=True)
                        elif 'So chuyen' in summary_df.columns:
                            summary_df = summary_df.sort_values('So chuyen', ascending=False).reset_index(drop=True)

                        if summary_df.empty:
                            st.info("Không có đủ dữ liệu để thống kê.")
                        else:
                            display_cols_spec = {"Tên Người dùng": 3}
                            if 'So chuyen' in summary_df.columns:
                                display_cols_spec["Số chuyến"] = 1
                            if 'Tong tien (VND)' in summary_df.columns:
                                display_cols_spec["Tổng tiền (VND)"] = 2

                            header_cols = st.columns(list(display_cols_spec.values()))
                            for i, col_name in enumerate(display_cols_spec.keys()):
                                header_cols[i].markdown(f"**{col_name}**")

                            current_employees = summary_df['Employee Name'].tolist()
                            expanded_state = st.session_state.setdefault("expanded_employees", {})
                            stale_keys = [name for name in expanded_state.keys() if name not in current_employees]
                            for name in stale_keys:
                                del expanded_state[name]
                            for name in current_employees:
                                expanded_state.setdefault(name, False)

                            def toggle_employee_expansion(employee_name: str) -> None:
                                state = st.session_state["expanded_employees"]
                                state[employee_name] = not state.get(employee_name, False)

                            date_cols = ['Date', 'Date of Trip', 'Trip Date', 'Ngày', 'Date & Time (GMT+7)']
                            date_col_name = find_col(df_merged, date_cols)
                            pickup_col_name = 'GEMINI_PICKUP_ADDRESS' if 'GEMINI_PICKUP_ADDRESS' in df_merged.columns else None
                            dropoff_col_name = 'GEMINI_DROPOFF_ADDRESS' if 'GEMINI_DROPOFF_ADDRESS' in df_merged.columns else None

                            for idx, row in summary_df.iterrows():
                                employee_name = row['Employee Name']
                                expanded = st.session_state["expanded_employees"][employee_name]
                                row_container = st.container()
                                with row_container:
                                    row_cols = st.columns(list(display_cols_spec.values()))
                                    with row_cols[0]:
                                        label = f"{'▼' if expanded else '►'} {employee_name}"
                                        st.button(
                                            label,
                                            key=f"summary_row_{idx}",
                                            use_container_width=True,
                                            on_click=toggle_employee_expansion,
                                            args=(employee_name,),
                                        )
                                    
                                    current_col_index = 1
                                    if 'So chuyen' in summary_df.columns:
                                        with row_cols[current_col_index]:
                                            st.write(int(row['So chuyen']))
                                        current_col_index += 1
                                    if 'Tong tien (VND)' in summary_df.columns:
                                        with row_cols[current_col_index]:
                                            st.write(f"{row['Tong tien (VND)']:,.0f}")

                                if st.session_state["expanded_employees"].get(employee_name):
                                    employee_df = df_merged[df_merged['Employee Name'] == employee_name]
                                    usage_date_col = 'NGAY_BOOKING' if 'NGAY_BOOKING' in employee_df.columns else date_col_name
                                    detail_cols = [
                                        'Employee Name', 'Booking ID', pickup_col_name, dropoff_col_name,
                                        'GEMINI_NGAY_HD_INVOICE', 'HINH_THUC_TT', 'TIEN_TRC_THUE',
                                        'TIEN_THUE8', 'TONG_TIEN', usage_date_col, 'SO_HOA_DON',
                                    ]
                                    final_cols = [col for col in detail_cols if col and col in employee_df.columns]

                                    detail_df = employee_df[final_cols].copy()
                                    detail_df.insert(0, "STT", range(1, len(detail_df) + 1))
                                    if 'SO_HOA_DON' in detail_df.columns:
                                        detail_df['SO_HOA_DON'] = detail_df['SO_HOA_DON'].astype(str).str.split('_').str[0].replace('nan', '')

                                    rename_map = {
                                        'Employee Name': 'Người sử dụng', 'Booking ID': 'Mã đặt chỗ',
                                        'GEMINI_PICKUP_ADDRESS': 'Điểm đón', 'GEMINI_DROPOFF_ADDRESS': 'Điểm đến',
                                        'GEMINI_NGAY_HD_INVOICE': 'Ngày HĐ', 'HINH_THUC_TT': 'Hình thức thanh toán',
                                        'TIEN_TRC_THUE': 'Tổng tiền trước thuế', 'TIEN_THUE8': 'Tổng tiền thuế (8%)',
                                        'TONG_TIEN': 'Tổng tiền đã có thuế', 'NGAY_BOOKING': 'Ngày sử dụng',
                                        'SO_HOA_DON': 'Số hóa đơn',
                                    }
                                    if usage_date_col and usage_date_col not in rename_map:
                                        rename_map[usage_date_col] = 'Ngày sử dụng'
                                    detail_df.rename(columns={k: v for k, v in rename_map.items() if k in detail_df.columns}, inplace=True)

                                    employee_display_col = rename_map.get('Employee Name', 'Employee Name')
                                    money_cols = [
                                        rename_map.get('TIEN_TRC_THUE', 'TIEN_TRC_THUE'),
                                        rename_map.get('TIEN_THUE8', 'TIEN_THUE8'),
                                        rename_map.get('TONG_TIEN', 'TONG_TIEN'),
                                    ]
                                    
                                    # Only add total row if there are money columns to sum
                                    if any(col in detail_df.columns for col in money_cols):
                                        total_row = {col: "" for col in detail_df.columns}
                                        total_row['STT'] = ""
                                        if employee_display_col in total_row:
                                            total_row[employee_display_col] = 'Tổng cộng'
                                        for money_col in money_cols:
                                            if money_col in detail_df.columns:
                                                numeric_values = pd.to_numeric(detail_df[money_col], errors='coerce')
                                                total_row[money_col] = numeric_values.sum()
                                        detail_df = pd.concat([detail_df, pd.DataFrame([total_row])], ignore_index=True)

                                    for money_col in money_cols:
                                        if money_col in detail_df.columns:
                                            numeric_series = pd.to_numeric(detail_df[money_col], errors='coerce')
                                            detail_df[money_col] = numeric_series.apply(
                                                lambda value: f"{value:,.0f}" if pd.notna(value) else ""
                                            )

                                    st.dataframe(detail_df, use_container_width=True, hide_index=True)

                                st.markdown('<hr style="margin-top:0.25rem; margin-bottom:0.25rem;">', unsafe_allow_html=True)
                try:
                    date_cols = ['Date', 'Date of Trip', 'Trip Date', 'Ngày', 'Date & Time (GMT+7)']
                    date_col_name = find_col(df_merged, date_cols)
                    pickup_col_name = 'GEMINI_PICKUP_ADDRESS' if 'GEMINI_PICKUP_ADDRESS' in df_merged.columns else None
                    dropoff_col_name = 'GEMINI_DROPOFF_ADDRESS' if 'GEMINI_DROPOFF_ADDRESS' in df_merged.columns else None
        
                    if date_col_name is None:
                        try:
                            uploaded_transport_file.seek(0)
                            header_list = pd.read_csv(uploaded_transport_file, skiprows=8, nrows=1).columns.tolist() if uploaded_transport_file.name.endswith('.csv') else pd.read_excel(uploaded_transport_file, skiprows=8, nrows=1).columns.tolist()
                            st.error(f"**Lỗi Bảng kê: Không thể tìm thấy cột Ngày tháng.** Vui lòng cho tôi biết tên cột ngày tháng chính xác từ danh sách bên dưới. **Các cột tìm được:** `{header_list}`")
                        except Exception as e:
                            st.error(f"Lỗi Bảng kê: Không tìm thấy cột ngày tháng. Lỗi khi đọc cột: {e}")
                    else:
                        df_merged['Date_dt'] = pd.to_datetime(df_merged[date_col_name], errors='coerce')
                        start_date = df_merged['Date_dt'].min()
                        end_date = df_merged['Date_dt'].max()
        
                        @st.cache_data
                        def generate_bang_ke_excel(_df, s_date, e_date, _d_col, _p_col, _do_col):
                            from openpyxl import load_workbook
                            from openpyxl.styles import Font
                            template_path = "FileMau/BangKe.xlsx"
                            wb = load_workbook(template_path)
                            ws = wb.active
                            ws['C4'] = s_date.strftime('%d/%m/%Y') if pd.notna(s_date) else "N/A"
                            ws['C5'] = e_date.strftime('%d/%m/%Y') if pd.notna(e_date) else "N/A"
                            start_row = 8
                            for i, row in _df.reset_index(drop=True).iterrows():
                                ws.cell(row=start_row + i, column=1, value=i + 1)
                                ws.cell(row=start_row + i, column=2, value=row['Booking ID'])
                                ws.cell(row=start_row + i, column=3, value=f" {row['Employee Name']}")
                                if _p_col: ws.cell(row=start_row + i, column=4, value=row[_p_col])
                                if _do_col: ws.cell(row=start_row + i, column=5, value=row[_do_col])
                                if 'GEMINI_NGAY_HD_INVOICE' in row and pd.notna(row['GEMINI_NGAY_HD_INVOICE']): ws.cell(row=start_row + i, column=6, value=row['GEMINI_NGAY_HD_INVOICE'])
                                ws.cell(row=start_row + i, column=7, value=row['HINH_THUC_TT'])
                                ws.cell(row=start_row + i, column=8, value="{:,.0f}".format(row['TIEN_TRC_THUE']))
                                ws.cell(row=start_row + i, column=9, value="{:,.0f}".format(row['TIEN_THUE8']))
                                ws.cell(row=start_row + i, column=10, value="{:,.0f}".format(row['TONG_TIEN']))
                                ws.cell(row=start_row + i, column=11, value=row['NGAY_BOOKING'])
                                ws.cell(row=start_row + i, column=12, value=row['pdf_link_key'])
                            total_row_index = start_row + len(_df)
                            total_label_cell = ws.cell(row=total_row_index, column=7, value="Tổng cộng"); total_label_cell.font = Font(bold=True)
                            total_value_cell_8 = ws.cell(row=total_row_index, column=8, value="{:,.0f}".format(_df['TIEN_TRC_THUE'].sum())); total_value_cell_8.font = Font(bold=True)
                            total_value_cell_9 = ws.cell(row=total_row_index, column=9, value="{:,.0f}".format(_df['TIEN_THUE8'].sum())); total_value_cell_9.font = Font(bold=True)
                            total_value_cell_10 = ws.cell(row=total_row_index, column=10, value="{:,.0f}".format(_df['TONG_TIEN'].sum())); total_value_cell_10.font = Font(bold=True)
                            excel_buffer = io.BytesIO(); wb.save(excel_buffer); return excel_buffer.getvalue()
        
                        st.subheader("📧 Gửi Bảng Kê qua Email (bằng Gmail)")
        
                        uploaded_email_mapping_file = st.file_uploader(
                            "Tải file Email Mapping (bắt buộc để gửi mail)",
                            type=["xlsx", "xls"],
                            help="Tải lên file Excel chứa cột 'Đơn vị' và 'Email' để gửi bảng kê."
                        )
        
                        # Read the mapping file as soon as it's uploaded
                        unit_to_email_map_upload = None
                        if uploaded_email_mapping_file is not None:
                            try:
                                uploaded_email_mapping_file.seek(0)
                                df_email_map_upload = pd.read_excel(uploaded_email_mapping_file)
                                email_col_upload = df_email_map_upload.columns[3]
                                unit_col_upload = df_email_map_upload.columns[4]
                                df_email_map_upload = df_email_map_upload.dropna(subset=[email_col_upload, unit_col_upload])
                                unit_to_email_map_upload = df_email_map_upload.groupby(unit_col_upload)[email_col_upload].apply(lambda x: list(x.unique())).to_dict()
                            except Exception as e:
                                st.error(f"Lỗi khi đọc file Email Mapping: {e}")
                                # Leave map as None and the error will be handled below
        
                        with st.expander("Hướng dẫn & Tải file mẫu Email Mapping"):
                            st.info("Để gửi email, bạn cần tải lên file Excel chứa thông tin email của các đơn vị. Bạn có thể tải file mẫu bên dưới để xem định dạng.")
                            try:
                                with open("FileMau/Tong hop _ Report.xlsx", "rb") as file:
                                    st.download_button(
                                        label="📥 Tải file mẫu (Tong hop _ Report.xlsx)",
                                        data=file,
                                        file_name="Tong hop _ Report.xlsx",
                                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                                    )
                            except FileNotFoundError:
                                st.error("Lỗi: Không tìm thấy file mẫu tại `FileMau/Tong hop _ Report.xlsx`.")
        
                        if 'credentials_json_content' not in st.session_state:
                            st.warning("Vui lòng tải file `credentials.json` ở trên để kích hoạt chức năng gửi email.")
                        else:
                            # --- SINGLE SEND ---
                            st.markdown("###### Gửi cho 1 đơn vị")
                            unit_col = 'Đơn vị'
                            unique_units_for_select_email = sorted(df_merged[unit_col].dropna().unique())
                            selected_unit_email = st.selectbox("Chọn đơn vị để gửi email", unique_units_for_select_email, key="email_unit_select")
        
                            if selected_unit_email:
                                # Display emails for the selected unit if the map is loaded
                                if uploaded_email_mapping_file and unit_to_email_map_upload is not None:
                                    recipient_emails_display = unit_to_email_map_upload.get(selected_unit_email, [])
                                    if recipient_emails_display:
                                        st.info(f"Bảng kê cho '{selected_unit_email}' sẽ được gửi đến các địa chỉ sau:")
                                        st.markdown(f"**{', '.join(recipient_emails_display)}**")
                                    else:
                                        st.warning(f"Không tìm thấy địa chỉ email cho đơn vị '{selected_unit_email}' trong file đã tải lên.")
        
                                if st.button(f"📧 Gửi Email đến '{selected_unit_email}'", use_container_width=True, key="send_email_btn"):
                                    if unit_to_email_map_upload is None:
                                        st.error("Vui lòng tải lên file Email Mapping hợp lệ trước khi gửi.")
                                    else:
                                        recipient_emails = unit_to_email_map_upload.get(selected_unit_email, [])
                                        if not recipient_emails:
                                            st.error(f"Không thể gửi email: Không tìm thấy email cho đơn vị '{selected_unit_email}' trong file đã tải lên.")
                                        else:
                                            to_field = ", ".join(recipient_emails)
                                            with st.spinner(f"Đang xác thực và gửi email đến {to_field}..."):
                                                try:
                                                    creds, _ = get_google_credentials(st.session_state.credentials_json_content)
                                                    if not creds:
                                                        st.error("Không thể lấy thông tin đăng nhập. Vui lòng thử đăng nhập lại.")
                                                        st.stop()
                                                    df_unit = df_merged[df_merged[unit_col] == selected_unit_email]
                                                    
                                                    # 1. Create Excel attachment
                                                    excel_data_email = generate_bang_ke_excel(df_unit, df_unit['Date_dt'].min(), df_unit['Date_dt'].max(), date_col_name, pickup_col_name, dropoff_col_name)
                                                    safe_unit_name = "".join(c for c in str(selected_unit_email) if c.isalnum() or c in (' ', '_')).rstrip()
                                                    excel_filename = f"BangKe_{safe_unit_name}.xlsx"
                                                    attachments = [{'data': excel_data_email, 'filename': excel_filename}]
        
                                                    # 2. Create zipped PDF attachments for each employee
                                                    if 'pdf_content' in df_unit.columns:
                                                        df_unit_with_pdfs = df_unit[df_unit['pdf_content'].notna()]
                                                        if not df_unit_with_pdfs.empty:
                                                            employee_groups = df_unit_with_pdfs.groupby('Employee Name')
                                                            st.info(f"Đang tạo và đính kèm {len(employee_groups)} file .zip (chứa PDF) cho từng nhân viên...")
                                                            for employee_name, employee_df in employee_groups:
                                                                zip_buffer = io.BytesIO()
                                                                with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_f:
                                                                    for _, row in employee_df.iterrows():
                                                                        if pd.notna(row['pdf_filename']) and pd.notna(row['pdf_content']):
                                                                            pdf_filename = os.path.basename(row['pdf_filename'])
                                                                            zip_f.writestr(pdf_filename, row['pdf_content'])
                                                                
                                                                safe_employee_name = "".join(c for c in str(employee_name) if c.isalnum() or c in (' ', '_')).rstrip()
                                                                zip_filename = f"{safe_employee_name}_pdf.zip"
                                                                
                                                                attachments.append({
                                                                    'data': zip_buffer.getvalue(),
                                                                    'filename': zip_filename
                                                                })
        
                                                    # 2.1. Create zipped XML attachments for each employee
                                                    if 'xml_content' in df_unit.columns:
                                                        df_unit_with_xmls = df_unit[df_unit['xml_content'].notna()]
                                                        if not df_unit_with_xmls.empty:
                                                            employee_groups_xml = df_unit_with_xmls.groupby('Employee Name')
                                                            st.info(f"Đang tạo và đính kèm {len(employee_groups_xml)} file .zip (chứa XML) cho từng nhân viên...")
                                                            for employee_name, employee_df in employee_groups_xml:
                                                                zip_buffer = io.BytesIO()
                                                                with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_f:
                                                                    for _, row in employee_df.iterrows():
                                                                        if pd.notna(row['xml_filename']) and pd.notna(row['xml_content']):
                                                                            xml_filename = os.path.basename(row['xml_filename'])
                                                                            zip_f.writestr(xml_filename, row['xml_content'])
                                                                
                                                                safe_employee_name = "".join(c for c in str(employee_name) if c.isalnum() or c in (' ', '_')).rstrip()
                                                                zip_filename = f"{safe_employee_name}_xml.zip"
                                                                
                                                                attachments.append({
                                                                    'data': zip_buffer.getvalue(),
                                                                    'filename': zip_filename
                                                                })
        
                                                    # 3. Send email
                                                    subject = f"HOA DON GRAP"
                                                    body = f"Kính gửi Cơ sở {selected_unit_email},\n\nTrung tâm xin gửi hóa đơn Grap phát sinh trong kỳ. Cán bộ thanh toán cơ sở vui lòng xem các file bảng kê và hóa đơn (nếu có) được đính kèm trong email này và thực hiện hồ sơ thanh toán đúng hạn.\n\nMọi thông tin thắc mắc, xin vui lòng liên hệ: lientt3@fe.edu.vn\nĐây là hệ thống đối chiếu tự động, vui lòng không reply email.\n\nTrân trọng"
                                                    send_gmail_message(creds, to_field, subject, body, attachments)
                                                    st.success(f"✅ Đã gửi email thành công đến {to_field} cho đơn vị '{selected_unit_email}'.")
                                                except Exception as e:
                                                    st.error(f"Lỗi khi gửi email: {e}")
        
                            # --- BULK SEND ---
                            st.divider()
                            st.markdown("###### Gửi cho tất cả các đơn vị")
                            if st.button("🚀 Gửi Email cho TẤT CẢ các đơn vị", use_container_width=True, key="send_all_emails_btn"):
                                if unit_to_email_map_upload is None:
                                    st.error("Vui lòng tải lên file Email Mapping hợp lệ trước khi gửi.")
                                else:
                                    with st.spinner("Bắt đầu quá trình gửi email hàng loạt..."):
                                        creds, _ = get_google_credentials(st.session_state.credentials_json_content)
                                        if not creds:
                                            st.error("Không thể lấy thông tin đăng nhập. Vui lòng thử đăng nhập lại.")
                                            st.stop()
                                        units_to_email = sorted(df_merged[unit_col].dropna().unique())
                                        progress_bar = st.progress(0, text="Bắt đầu...")
                                        success_count = 0
                                        failed_units = []
                                        for i, unit in enumerate(units_to_email):
                                            progress_text = f"Đang xử lý: {unit} ({i+1}/{len(units_to_email)})"; progress_bar.progress((i + 1) / len(units_to_email), text=progress_text)
                                            recipient_emails = unit_to_email_map_upload.get(unit, [])
                                            if not recipient_emails:
                                                failed_units.append((unit, "Không tìm thấy email trong file mapping đã tải lên."))
                                                continue
                                            try:
                                                df_unit = df_merged[df_merged[unit_col] == unit]
                                                
                                                # 1. Create Excel attachment
                                                excel_data_email = generate_bang_ke_excel(df_unit, df_unit['Date_dt'].min(), df_unit['Date_dt'].max(), date_col_name, pickup_col_name, dropoff_col_name)
                                                safe_unit_name = "".join(c for c in str(unit) if c.isalnum() or c in (' ', '_')).rstrip()
                                                excel_filename = f"BangKe_{safe_unit_name}.xlsx"
                                                attachments = [{'data': excel_data_email, 'filename': excel_filename}]
        
                                                # 2. Create zipped PDF attachments for each employee in the unit
                                                if 'pdf_content' in df_unit.columns:
                                                    df_unit_with_pdfs = df_unit[df_unit['pdf_content'].notna()]
                                                    if not df_unit_with_pdfs.empty:
                                                        for employee_name, employee_df in df_unit_with_pdfs.groupby('Employee Name'):
                                                            zip_buffer = io.BytesIO()
                                                            with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_f:
                                                                for _, row in employee_df.iterrows():
                                                                    if pd.notna(row['pdf_filename']) and pd.notna(row['pdf_content']):
                                                                        pdf_filename = os.path.basename(row['pdf_filename'])
                                                                        zip_f.writestr(pdf_filename, row['pdf_content'])
                                                            
                                                            safe_employee_name = "".join(c for c in str(employee_name) if c.isalnum() or c in (' ', '_')).rstrip()
                                                            zip_filename = f"{safe_employee_name}_pdf.zip"
                                                            
                                                            attachments.append({
                                                                'data': zip_buffer.getvalue(),
                                                                'filename': zip_filename
                                                            })
        
                                                # 2.1. Create zipped XML attachments for each employee
                                                if 'xml_content' in df_unit.columns:
                                                    df_unit_with_xmls = df_unit[df_unit['xml_content'].notna()]
                                                    if not df_unit_with_xmls.empty:
                                                        for employee_name, employee_df in df_unit_with_xmls.groupby('Employee Name'):
                                                            zip_buffer = io.BytesIO()
                                                            with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_f:
                                                                for _, row in employee_df.iterrows():
                                                                    if pd.notna(row['xml_filename']) and pd.notna(row['xml_content']):
                                                                        xml_filename = os.path.basename(row['xml_filename'])
                                                                        zip_f.writestr(xml_filename, row['xml_content'])
                                                            
                                                            safe_employee_name = "".join(c for c in str(employee_name) if c.isalnum() or c in (' ', '_')).rstrip()
                                                            zip_filename = f"{safe_employee_name}_xml.zip"
                                                            
                                                            attachments.append({
                                                                'data': zip_buffer.getvalue(),
                                                                'filename': zip_filename
                                                            })
        
                                                # 3. Send email
                                                subject = f"Bảng kê đối chiếu Grab cho đơn vị '{unit}'"
                                                body = f"Kính gửi Quý đơn vị {unit},\n\nVui lòng xem các file bảng kê và hóa đơn (nếu có) được đính kèm trong email này.\n\nTrân trọng,\nSerder mail."
                                                to_field = ", ".join(recipient_emails)
                                                send_gmail_message(creds, to_field, subject, body, attachments)
                                                success_count += 1
                                            except Exception as e:
                                                failed_units.append((unit, str(e)))
                                        progress_bar.empty()
                                        st.success(f"✅ Hoàn tất! Đã gửi thành công {success_count}/{len(units_to_email)} email.")
                                        if failed_units:
                                            st.error(f"❌ Có {len(failed_units)} email gửi thất bại.")
                                            with st.expander("Xem chi tiết lỗi"):
                                                for unit, reason in failed_units: st.write(f"- **{unit}**: {reason}")
                except Exception as e:
                    st.error(f"Đã xảy ra lỗi khi tạo file Bảng kê hoặc gửi mail: {e}")

        except Exception as e:
            st.error(f"Đã xảy ra lỗi trong quá trình xử lý: {e}")
            st.exception(e)

# --- HÀM HỖ TRỢ ---
def find_col(df, possibilities):
    """Finds the first column in a dataframe that exists from a list of possibilities."""
    for p in possibilities:
        if p in df.columns:
            return p
    return None

def load_mapping_data():
    """Đọc file Excel mapping và trả về 2 dictionaries:
    1. Employee Name -> Đơn vị
    2. Đơn vị -> List of Emails
    """
    try:
        df_mapping = pd.read_excel("FileMau/Tong hop _ Report.xlsx")
        name_col = df_mapping.columns[1]
        email_col = df_mapping.columns[3]
        unit_col = df_mapping.columns[4]
        df_mapping = df_mapping.dropna(subset=[name_col, unit_col])
        employee_to_unit_map = df_mapping.set_index(name_col)[unit_col].to_dict()
        df_email_map = df_mapping.dropna(subset=[email_col])
        # Group by unit and create a list of unique emails for each unit
        unit_to_email_map = df_email_map.groupby(unit_col)[email_col].apply(lambda x: list(x.unique())).to_dict()
        return employee_to_unit_map, unit_to_email_map
    except FileNotFoundError:
        st.error("Lỗi: Không tìm thấy file mapping 'FileMau/Tong hop _ Report.xlsx'. Vui lòng đảm bảo file tồn tại.")
        st.stop()
    except IndexError:
        st.error("Lỗi: File mapping 'Tong hop _ Report.xlsx' không có đủ 5 cột (để lấy cột B, D, và E).")
        st.stop()
    except Exception as e:
        st.error(f"Lỗi khi đọc file mapping: {e}")
        st.stop()

def send_gmail_message(credentials, to, subject, body, attachments=None):
    """Sends an email with multiple attachments using Gmail API."""
    try:
        # Lấy thông tin credentials từ session state
        creds_json = json.loads(st.session_state['credentials'])
        creds = Credentials.from_authorized_user_info(creds_json, SCOPES)
        
        service = build('gmail', 'v1', credentials=creds)
        user_profile = service.users().getProfile(userId='me').execute()
        sender_email = user_profile['emailAddress']
        sender_name = "Hệ thống đối chiếu tự động"
        
        message = MIMEMultipart()
        message['to'] = to
        message['from'] = formataddr((sender_name, sender_email))
        message['subject'] = subject
        msg_body = MIMEText(body, 'plain', 'utf-8')
        message.attach(msg_body)

        if attachments:
            for attachment in attachments:
                if attachment and attachment.get('data') and attachment.get('filename'):
                    part = MIMEApplication(attachment['data'], Name=attachment['filename'])
                    part['Content-Disposition'] = f'attachment; filename="{attachment["filename"]}"'
                    message.attach(part)

        encoded_message = base64.urlsafe_b64encode(message.as_bytes()).decode()
        create_message = {'raw': encoded_message}
        send_message = (service.users().messages().send(userId="me", body=create_message).execute())
    except HttpError as error:
        st.error(f"An error occurred while sending email: {error}")
        raise error
    except KeyError:
        st.error("Lỗi: Không tìm thấy thông tin đăng nhập trong session. Vui lòng đăng nhập lại.")
        st.stop()

# --- ĐIỂM BẮT ĐẦU CỦA APP ---
if 'user_info' in st.session_state:
    main_app()
else:
    show_login_page()